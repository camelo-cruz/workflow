#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Copyright (C) 2024  Alejandra Camelo Cruz, Arne Goelz

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.
...
Leibniz Institute General Linguistics (ZAS)
"""

import os
import re
import warnings
import logging
import sys
import argparse
import pandas as pd
import openpyxl
from dotenv import load_dotenv 
from openpyxl.styles import Font
from tqdm import tqdm
import whisper
import whisperx
import torch

from functions import set_global_variables, find_language, clean_string, find_ffmpeg

# Set global variables and suppress warnings
LANGUAGES, NO_LATIN, OBLIGATORY_COLUMNS, _ = set_global_variables()
warnings.filterwarnings("ignore")

# Configure global logger
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.WARNING)
console_handler.setFormatter(logging.Formatter("%(message)s"))
logger.addHandler(console_handler)

ffmpeg_path = find_ffmpeg()


class Transcriber:
    def __init__(self, input_dir, language, device=None):
        self.input_dir = input_dir
        self.language_code = find_language(language, LANGUAGES)
        self.device = device if device is not None else ("cuda" if torch.cuda.is_available() else "cpu")
        self.batch_size = 16 # reduce if low on GPU mem
        try:
            self.model = whisperx.load_model("large-v2", device, compute_type="float16", language=self.language_code)
        except:
            self.model = whisperx.load_model("large-v2", device, compute_type="int8", language=self.language_code)
        
        ## Load Hugging
        try:
            # If running under PyInstaller, sys._MEIPASS is available
            base_path = os.path.join(sys._MEIPASS, 'materials')
            print("Using sys._MEIPASS for materials path")
        except Exception:
            # Fallback to using the script's directory if not running as a PyInstaller bundle
            script_dir = os.path.dirname(os.path.abspath(__file__))
            base_path = os.path.join(script_dir, 'materials')
            print("Using script directory for materials path")

        secrets_path = os.path.join(base_path, 'secrets.env')

        if os.path.exists(secrets_path):
            load_dotenv(secrets_path, override=True)
        else:
            print(f"Error: {secrets_path} not found.")
            sys.exit(1)
        
        self.hugging_key = os.getenv("HUGGING_KEY")
        if not self.hugging_key:
            raise ValueError("Hugging face key not found. Check your secrets.env file.")
        

    def setup_logging(self, log_file_path):
        """Set up file logging for a given directory."""
        file_handler = logging.FileHandler(log_file_path)
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(
            logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
        )
        logger.addHandler(file_handler)
        logger.info(f"Logging to {log_file_path}")
        logger.info(f"Using ffmpeg from {ffmpeg_path}")
        return file_handler

    def load_trials_data(self, base_dir):
        """
        Load the trials and sessions data from CSV or Excel.
        Returns the dataframe and the path for the annotated Excel output.
        """
        csv_file = os.path.join(base_dir, 'trials_and_sessions.csv')
        excel_file = os.path.join(base_dir, 'trials_and_sessions.xlsx')
        excel_out = os.path.join(base_dir, 'trials_and_sessions_annotated.xlsx')

        if os.path.exists(csv_file):
            df = pd.read_csv(csv_file)
        elif os.path.exists(excel_file):
            df = pd.read_excel(excel_file)
        else:
            raise FileNotFoundError("No trials_and_sessions file found in the directory.")

        # Ensure obligatory columns exist
        for col in OBLIGATORY_COLUMNS:
            if col not in df:
                df[col] = ""
        if self.language_code not in NO_LATIN:
            df["transcription_original_script"] = ""
            df["transcription_original_script_utterance_used"] = ""

        return df, excel_out

    def _append_to_cell(self, df, idx, column, text):
        """Helper to append text to a DataFrame cell, initializing if empty."""
        old_val = df.at[idx, column]
        df.at[idx, column] = ("" if pd.isna(old_val) else old_val) + text

    def add_transcription_to_df(self, df, file, transcription, count, filename_regexp):
        """
        Add the transcription text to the dataframe.
        If the file is not found, use block/task/trial pattern to locate the row.
        """
        # locate any direct match
        series = df[df.isin([file])].stack()
        is_nonlatin = self.language_code in NO_LATIN
        text_auto = f"{count}: {transcription}"
        text_suffix = " - " if series.empty else " "
        col_name = (
            'transcription_original_script' if is_nonlatin
            else 'latin_transcription_everything'
        )

        if series.empty:
            match = filename_regexp.search(file)
            if not match:
                logger.warning(
                    f"File '{file}' does not match block/task/trial pattern. Skipping."
                )
                return

            blk = int(match.group('block'))
            tsk = int(match.group('task'))
            trl = int(match.group('trial'))
            cond = (
                (df['Block_Nr'] == blk) &
                (df['Task_Nr'] == tsk) &
                (df['Trial_Nr'] == trl)
            )
            if df.loc[cond].empty:
                logger.warning(
                    f"No row for block {blk}, task {tsk}, trial {trl}. Skipping '{file}'."
                )
                return

            # assign missing_filename to the first available slot
            col_ctr = 1
            miss_col = f'missing_filename_{col_ctr}'
            while (
                miss_col in df.columns and
                not df.loc[cond, miss_col].isna().all()
            ):
                col_ctr += 1
                miss_col = f'missing_filename_{col_ctr}'
            df.loc[cond, miss_col] = file

            # append transcription to each matched index
            for idx in df.loc[cond].index:
                self._append_to_cell(df, idx, 'automatic_transcription', text_auto + text_suffix)
                self._append_to_cell(df, idx, col_name, text_auto + text_suffix)

        else:
            # direct match: stack() yields (row_index, col_index) tuples
            for (row_idx, _col), _ in series.items():
                self._append_to_cell(df, row_idx, 'automatic_transcription', text_auto + text_suffix)
                self._append_to_cell(df, row_idx, col_name, text_auto + text_suffix)
    
    def transcribe_and_diarize(self, path_to_audio, prompt=''):
        # --- Chinese-only path: just transcribe and return raw text ---
        if self.language_code == 'zh':
            res = self.model.transcribe(
                path_to_audio,
                language='zh',
                initial_prompt="请使用简体中文转录。"
            )
            # strip out our prompt prefix, then return the result immediately
            clean = res["text"].replace("请使用简体中文转录。", "").strip()
            return clean

        # --- multilingual path: load, transcribe, align, diarize, then stitch back together ---
        else:
            audio = whisperx.load_audio(path_to_audio)

            result = self.model.transcribe(
                audio,
                batch_size=self.batch_size,
                language=self.language_code
            )

            # align word‐level timestamps
            model_a, metadata = whisperx.load_align_model(
                language_code=result["language"],
                device=self.device
            )
            result = whisperx.align(
                result["segments"],
                model_a,
                metadata,
                audio,
                self.device,
                return_char_alignments=False
            )

            # run diarization
            diarize_model = whisperx.DiarizationPipeline(
                use_auth_token=self.hugging_key,
                device=self.device)
            diarize_segments = diarize_model(audio)

            result = whisperx.assign_word_speakers(diarize_segments, result)

            full_sentences = []
            buffer_speaker = None
            buffer_text = ""

            for seg in result["segments"]:
                # 1) pick up the speaker if present, else reuse the last one
                spk = seg.get("speaker", buffer_speaker)
                
                # 2) if we still have no speaker (i.e. first segment was unlabeled), skip
                if spk is None:
                    continue
                
                txt = seg["text"].strip()
                
                # 3) start the first buffer
                if buffer_speaker is None:
                    buffer_speaker, buffer_text = spk, txt
                
                # 4) same speaker → just append
                elif spk == buffer_speaker:
                    buffer_text += " " + txt
                
                # 5) speaker changed → flush old and start new
                else:
                    full_sentences.append(f"{buffer_speaker}: {buffer_text}")
                    buffer_speaker, buffer_text = spk, txt

            # 6) flush whatever’s left
            if buffer_speaker is not None:
                full_sentences.append(f"{buffer_speaker}: {buffer_text}")

            return "  ".join(full_sentences)

                        
    def format_excel_output(self, excel_output_file):
        """Apply red font to target columns in the Excel output."""
        wb = openpyxl.load_workbook(excel_output_file)
        ws = wb.active
        red = Font(color="FF0000")
        targets = ['transcription_original_script', 'latin_transcription_everything']

        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        idx_map = {h: i+1 for i, h in enumerate(headers) if h in targets}

        for row in ws.iter_rows(min_row=2):
            for col, col_i in idx_map.items():
                cell = row[col_i-1]
                if cell.value:
                    cell.font = red

        wb.save(excel_output_file)
        logger.info(f"Excel saved and formatted: '{excel_output_file}'")

    def process_data(self, verbose=True):
        """Walk input directory, process audio, and update trials file."""
        filename_regexp = re.compile(
            r'blockNr_(?P<block>\d+)_taskNr_(?P<task>\d+)_trialNr_(?P<trial>\d+).*'
        )

        for subdir, _, files in os.walk(self.input_dir):
            if 'binaries' not in subdir:
                continue

            logger.info(f"Device: {self.model.device}")
            logger.info(f"Processing {subdir}")
            print(f"Processing {subdir}")
            base = os.path.abspath(os.path.join(subdir, '..'))
            log_path = os.path.join(base, "transcription.log")
            fh = self.setup_logging(log_path)

            try:
                df, out_file = self.load_trials_data(base)
            except FileNotFoundError as e:
                logger.error(e)
                logger.removeHandler(fh)
                fh.close()
                continue

            count = 0
            files.sort()
            for file in tqdm(files, desc="Transcribing"):
                if not file.lower().endswith(('.mp3', '.mp4', '.m4a')):
                    continue
                count += 1
                path = os.path.abspath(os.path.join(subdir, file))
                logger.debug(f"File {count}/{len(files)}: {path}")
                try:
                    text = self.transcribe_and_diarize(path)
                    text = clean_string(text)
                    if verbose:
                        tqdm.write(text)
                    self.add_transcription_to_df(df, file, text, count, filename_regexp)
                except Exception as e:
                    logger.error(f"Error on '{file}': {e}")
                    continue
            print('*************output inside transcriber:', out_file)
            df.to_excel(out_file, index=False)
            self.format_excel_output(out_file)
            logger.info(f"Completed '{subdir}'")
            logger.removeHandler(fh)
            fh.close()


def main():
    parser = argparse.ArgumentParser(description="Automatic transcription")
    parser.add_argument("input_dir", help="Directory containing audio files")
    parser.add_argument("language", help="Language of the audio content")
    parser.add_argument(
        "--verbose", action="store_true",
        help="Print detailed transcription output"
    )
    args = parser.parse_args()

    transcriber = Transcriber(args.input_dir, args.language)
    transcriber.process_data(verbose=args.verbose)


if __name__ == "__main__":
    main()