
# -*- coding: utf-8 -*-
"""
Copyright (C) 2025  Alejandra Camelo Cruz

    This program is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with this program.  If not, see <https://www.gnu.org/licenses/>.
        
Leibniz Institute General Linguistics (ZAS)
"""

import os
import argparse
import openpyxl
from openpyxl.styles import Font
import pandas as pd
from tqdm import tqdm
from functions import set_global_variables, find_language
from langchain_core.prompts import ChatPromptTemplate
from langchain_ollama.llms import OllamaLLM


LANGUAGES, NO_LATIN, OBLIGATORY_COLUMNS, _ = set_global_variables()

class SentenceSelector():
    def __init__(self, input_dir, language, study):
        self.input_dir = input_dir
        self.language = language
        self.language_code = find_language(language, LANGUAGES)
        self.study = study
        self.llm_model = OllamaLLM(model="mistral")

    def choose_sentences(self, df, verbose=False):
        if self.language_code in NO_LATIN:
            raise NotImplementedError("Function for no Latin languages not implemented yet.")
        else:
            source = "latin_transcription_everything"
            target = "latin_transcription_utterance_used"

        instruction = "relative clauses" if 'H' in self.study else "default instruction"

        # Define LangChain prompt template
        template = (
            f"This is a study about {instruction} in {self.language}. "
            f"In the following text, extract only the phrases or sentences that contain {instruction}. "
            f"Only select and list the phrases or sentences that contain such {instruction}. "
            "Keep the original order in which these appear in the text. "
            "Don't add anything else like e.g. commentary or translations to your response. "
            "Your response should only consist of the extracted sentences."
            "Here is the text: {text}"
        )
        prompt = ChatPromptTemplate.from_template(template)
        chain = prompt | self.llm_model

        df[target] = ""
        df[target] = df[target].astype('object')

        for text in df[source].dropna():
            series = df[df.isin([text])].stack()
            for idx, _ in series.items():
                if pd.isna(df.at[idx[0], target]):
                    df.at[idx[0], target] = ""

                response = chain.invoke({"text": text})
                response_text = str(response).strip()

                if verbose:
                    print(f"Response: {response_text}")

                if response_text:
                    df.at[idx[0], target] += response_text + "\n"

        return df

    def process_data(self, verbose=False):
        files_to_process = []
        for subdir, _, files in os.walk(self.input_dir):
            for file in files:
                if file.endswith('annotated.xlsx'):
                    files_to_process.append(os.path.join(subdir, file))

        # Process each file individually
        for file_path in tqdm(files_to_process, desc="Processing Files", unit="file"):
            print(f"Processing {file_path}...")
            df = pd.read_excel(file_path)
            df = self.choose_sentences(df, verbose=verbose)
            df.to_excel(file_path, index=False)

            # Open the Excel file and modify cell formatting
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active  # Select the first sheet

            # Define the red font style
            red_font = Font(color="FF0000")  # Hex code for red color


            target_column = 'latin_transcription_utterance_used'

            # Get the header row (first row) as a tuple of column names
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

            # Ensure the target column exists
            if target_column not in header_row:
                raise ValueError(f"Target column '{target_column}' not found in header row.")

            # Find the column index (1-based) for the target column
            target_index = header_row.index(target_column) + 1

            # Apply red font to each cell in the target column (skip header row)
            for row in ws.iter_rows(min_row=2):
                cell = row[target_index - 1]  # Adjust for 0-based index in row
                if cell.value:
                    cell.font = red_font

            # Save the modified Excel file
            wb.save(file_path)


        print("Processing completed.")


def main():
    parser = argparse.ArgumentParser(description="Automatic selection of sentences.")
    parser.add_argument("input_dir", help="Directory with files to select.")
    parser.add_argument("source_language", help="Source language for selection.")
    parser.add_argument("study", help="Name of the study.")
    args = parser.parse_args()


    language = args.source_language.lower()

    # Create an instance of Transliterator and process the data
    sentenceSelector = SentenceSelector(args.input_dir, language, args.study)
    sentenceSelector.process_data()


if __name__ == "__main__":
    main()
