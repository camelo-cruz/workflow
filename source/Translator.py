# -*- coding: utf-8 -*-
"""
Copyright (C) 2024  Alejandra Camelo Cruz, Arne Goelz

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.

Leibniz Institute General Linguistics (ZAS)
"""

import os
import pandas as pd
import argparse
import torch
import deepl
import logging
import time
from tqdm import tqdm
import openpyxl
from dotenv import load_dotenv
from openpyxl.styles import Font
from functions import set_global_variables, find_language
from transformers import AutoTokenizer, AutoModelForSeq2SeqLM

def setup_logging(log_file_path):
    """ Dynamically updates logging to write to a new log file. """
    logger = logging.getLogger()
    
    # Remove existing handlers to avoid duplicate logs
    if logger.hasHandlers():
        logger.handlers.clear()

    # Create a file handler for the new log file
    file_handler = logging.FileHandler(log_file_path, mode="a")  # "a" = append mode
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))

    # Add console logging as well
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))

    # Add handlers to the logger
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    logging.info(f"Logging started for {log_file_path}")

LANGUAGES, NO_LATIN, OBLIGATORY_COLUMNS, _ = set_global_variables()


class Translator():
    def __init__(self, input_dir, language, instruction, device="cpu"):
        self.input_dir = input_dir
        self.language_code = find_language(language, LANGUAGES)
        self.instruction = instruction
        self.device = device

        self.tokenizer = AutoTokenizer.from_pretrained("facebook/nllb-200-1.3B", src_lang=f'{self.language_code}_Latn')
        logging.info(f"Initialized Translator for language: {language} (code: {self.language_code}) (instruction: {instruction})")

    def translate_with_pretrained(self, text):
        """ Translates text using the NLLB-200 model """
        start_time = time.time()
        self.model = AutoModelForSeq2SeqLM.from_pretrained("facebook/nllb-200-1.3B").to(self.device)
        inputs = self.tokenizer(text, return_tensors="pt").to(self.device)
        translated_tokens = self.model.generate(
            **inputs, forced_bos_token_id=self.tokenizer.convert_tokens_to_ids("eng_Latn")
        )

        translated = self.tokenizer.batch_decode(translated_tokens, skip_special_tokens=True)[0]
        end_time = time.time()

        logging.info(f"Translated with NLLB-200 in {end_time - start_time:.2f} seconds")
        return translated

    def translate_with_deepl(self, text):
        """ Translates text using DeepL API, ensuring correct language codes """
        load_dotenv("secrets.env")
        api_key = os.getenv("API_KEY")
        if not api_key:
            raise ValueError("API key not found. Check your secrets.env file.")

        start_time = time.time()
        deepl_client = deepl.DeepLClient(api_key)
        
        # Handle "PT" separately
        if self.language_code.lower() == "pt":
            self.language_code = "PT-BR"  # Default to Brazilian Portuguese

        source_lang = self.language_code.upper()  # Ensure uppercase

        try:
            result = deepl_client.translate_text(text, source_lang=source_lang, target_lang='EN-US')
        except deepl.DeepLException:
            result = deepl_client.translate_text(text, target_lang='EN-US')
        
        end_time = time.time()
        logging.info(f"Translated with DeepL in {end_time - start_time:.2f} seconds (Auto-detected language)")

        return result.text

    def process_data(self, verbose=False):
        """ Processes and translates data from input Excel files """
        logging.info(f"Starting translation process for directory: {self.input_dir}")
        start_time = time.time()

        automatic_column = "automatic_transcription"
        corrected_column = "latin_transcription_everything"
        sentences_column = "latin_transcription_utterance_used"

        if self.language_code in NO_LATIN:
            corrected_column = "transcription_original_script"
            sentences_column = "transcription_original_script_utterance_used"

        files_to_process = []
        for subdir, dirs, files in os.walk(self.input_dir):
            for file in files:
                if file.endswith('annotated.xlsx'):
                    files_to_process.append(os.path.join(subdir, file))
 
        with tqdm(total=len(files_to_process), desc="Processing files", unit="file") as file_pbar:
            for file_path in files_to_process:
                log_file = os.path.join(os.path.dirname(file_path), "translation.log")
                setup_logging(log_file)
                logging.info(f"Processing file: {file_path}")
                df = pd.read_excel(file_path)

                for i in range(len(df)):
                    print("translating row: ", i)
                    try:
                        text_to_translate = df.at[i, corrected_column if self.instruction == 'corrected' else
                                                        automatic_column if self.instruction == 'automatic' else
                                                        sentences_column]

                        if pd.isna(text_to_translate) or not str(text_to_translate).strip():
                            continue  # Skip empty values

                        translation = self.translate_with_pretrained(text_to_translate)
                        #ranslation = self.translate_with_deepl(text_to_translate)

                        if self.instruction == 'corrected':
                            df.at[i, "automatic_translation_corrected_transcription"] = translation
                            df.at[i, "translation_everything"] = translation
                        elif self.instruction == 'automatic':
                            df.at[i, "automatic_translation_automatic_transcription"] = translation
                        elif self.instruction == 'sentences':
                            df.at[i, "automatic_translation_utterance_used"] = translation
                            df.at[i, "translation_utterance_used"] = translation

                        if verbose:
                            print(f"Original: {text_to_translate}, Translation: {translation}")

                    except Exception as e:
                        logging.error(f"Error in row {i} of file {file_path}: {e}")

                # Reorder columns to ensure obligatory columns are at the end
                extra_columns = [col for col in df.columns if col not in OBLIGATORY_COLUMNS]
                df = df[extra_columns + [col for col in OBLIGATORY_COLUMNS if col in df.columns]]

                df.to_excel(file_path, index=False)

                # Open the Excel file and modify cell formatting
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active  # Select the first sheet

                # Define the red font style
                red_font = Font(color="FF0000")  # Hex code for red color

                # Define column names that should have red font
                target_columns = ['translation_everything', 'translation_utterance_used']

                # Find column indexes for the target column names
                header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))  # Get first row as header
                column_indexes = {col_name: idx + 1 for idx, col_name in enumerate(header_row) if col_name in target_columns}

                # Apply red font to the specified columns
                for row in ws.iter_rows(min_row=2):  # Skip header row
                    for col_name, col_idx in column_indexes.items():
                        cell = row[col_idx - 1]  # Column index is 1-based, row[] is 0-based
                        if cell.value:
                            cell.font = red_font

                # Save the modified Excel file
                wb.save(file_path)

                file_pbar.update(1)

        end_time = time.time()
        logging.info(f"Translation process completed in {end_time - start_time:.2f} seconds")


def main():
    """ Main function to translate manually prepared transcriptions """
    parser = argparse.ArgumentParser(description="automatic transcription")
    parser.add_argument("input_dir")
    parser.add_argument("language")
    parser.add_argument("--instruction", "-i",
                        choices=["automatic_transcription",
                                 "corrected_transcription",
                                 "sentences"],
                        help="Type of instruction for translation", required=False)
    args = parser.parse_args()

    translator = Translator(args.input_dir, args.language, args.instruction)
    translator.process_data()


if __name__ == "__main__":
    main() 
