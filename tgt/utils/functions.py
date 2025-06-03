import os
import sys
import json
import string
import os
import shutil
import logging
import subprocess
import urllib.request
import zipfile
import openpyxl

from openpyxl.styles import Font

def load_json_file(file_path):
    """Utility function to load JSON files with error handling."""
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return json.load(file)
    except FileNotFoundError:
        print(f"Error: {file_path} not found.")
        sys.exit(1)
    except json.JSONDecodeError:
        print(f"Error: Failed to parse JSON from {file_path}.")
        sys.exit(1)

def get_materials_path(filename):
    """Get the path to a file in the materials directory."""
    parent_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        base_path = os.path.join(sys._MEIPASS, parent_dir, 'materials')
    except:
        base_path = os.path.join(os.getcwd(), parent_dir, 'materials')

    return os.path.join(base_path, filename)

def load_text_file(filename):
    """Utility function to load text files with error handling."""
    # Resolve the file path using get_materials_path
    file_path = get_materials_path(filename)
    
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read().splitlines()
    except FileNotFoundError:
        print(f"Error: {file_path} not found.")
        sys.exit(1)

def set_global_variables():
    """Loads necessary configurations and paths."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)
    
    languages_path = os.path.join(parent_dir, 'materials', 'LANGUAGES')
    columns_path = os.path.join(parent_dir, 'materials', 'OBLIGATORY_COLUMNS')
    nolatin_path = os.path.join(parent_dir, 'materials', 'NO_LATIN')

    LANGUAGES = load_json_file(languages_path)
    OBLIGATORY_COLUMNS = load_text_file(columns_path)
    NO_LATIN = load_text_file(nolatin_path)

    return LANGUAGES, NO_LATIN, OBLIGATORY_COLUMNS

def load_glossing_rules(filename):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    parent_dir = os.path.dirname(script_dir)

    file = os.path.join(parent_dir, 'materials', 'glossing_rules', filename)

    return load_json_file(file)

def find_language(language, LANGUAGES):
    """Finds the language code by its name."""
    language_lower = language.lower()
    
    # Reverse the LANGUAGES dictionary for language name -> code lookup
    reversed_languages = {value.lower(): key for key, value in LANGUAGES.items()}
    
    language_code = reversed_languages.get(language_lower)
    
    if language_code:
        print(f'Language recognized: {language_code}')
        return language_code
    else:
        print(f"Unsupported language: {language}")
        sys.exit(1)

def clean_string(input_string):
    """
    Process a string by converting it to lowercase and removing punctuation.

    Parameters:
        input_string (str): The string to be processed.

    Returns:
        str: The processed string.
    """
    lowercase_string = input_string.lower()

    translator = str.maketrans("", "", string.punctuation)
    processed_string = lowercase_string.translate(translator)

    return processed_string


def install_ffmpeg():
    """
    Downloads FFmpeg, unpacks zip-file, deletes zip-path
    and installs FFmpeg to specific path
    """
    destination_path = os.path.expanduser("~")
    ffmpeg_url = "https://www.gyan.dev/ffmpeg/builds/ffmpeg-release-essentials.zip"

    zip_path = os.path.join(destination_path, "ffmpeg-7.1-essentials_build.zip")
    ffmpeg_extract_path = os.path.join(destination_path, "ffmpeg")

    try:
        print("Downloading ffmpeg...")
        urllib.request.urlretrieve(ffmpeg_url, zip_path)
        print("Download complete.")

        print(f"Extracting ffmpeg to {ffmpeg_extract_path}...")
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(ffmpeg_extract_path)
        print("Extraction complete.")

        os.remove(zip_path)

        print(f"ffmpeg has been installed to {ffmpeg_extract_path}.")
        print("Adding path to system's PATH environment variable.")
        ffmpeg_path = os.path.join(ffmpeg_extract_path, "ffmpeg-7.1-essentials_build/bin/ffmpeg.exe")
        os.environ["PATH"] += os.pathsep + os.path.dirname(ffmpeg_path)
        return ffmpeg_path
    except Exception as e:
        print("An error occurred:", e)


def find_ffmpeg():
    """Dynamically finds ffmpeg executable path"""
    ffmpeg_path = shutil.which("ffmpeg")

    if not ffmpeg_path:
        print("FFmpeg not found. Attempting to install FFmpeg...")
        ffmpeg_path = install_ffmpeg()
    else:
        return ffmpeg_path


def format_excel_output(excel_output_file, columns_to_highlight: list):
    wb = openpyxl.load_workbook(excel_output_file)
    ws = wb.active
    red = Font(color="FF0000")
    headers = [cell.value for cell in ws[1]]
    idx_map = {h: i+1 for i,h in enumerate(headers) if h in columns_to_highlight}
    for row in ws.iter_rows(min_row=2):
        for col, col_i in idx_map.items():
            cell = row[col_i-1]
            if cell.value:
                cell.font = red
    wb.save(excel_output_file)


def setup_logging(logger, log_path):
    logger.setLevel(logging.DEBUG)

    # Clear existing handlers (avoid duplicates if run multiple times)
    if logger.hasHandlers():
        logger.handlers.clear()

    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.DEBUG)
    console_formatter = logging.Formatter("%(message)s")
    console_handler.setFormatter(console_formatter)

    # File handler
    file_handler = logging.FileHandler(log_path, mode='w', encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    file_handler.setFormatter(file_formatter)

    # Add both handlers
    logger.addHandler(console_handler)
    logger.addHandler(file_handler)

    return file_handler  # so you can later remove it if needed

