# -*- coding: utf-8 -*-
"""
Copyright (C) 2024  Alejandra Camelo Cruz, Arne Goelz

    This program is free software: you can redistribute it and/or modify
    it under the terms of the GNU General Public License as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    This program is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU General Public License for more details.

    You should have received a copy of the GNU General Public License
    along with this program.  If not, see <https://www.gnu.org/licenses/>.
        
Leibniz Institute General Linguistics (ZAS)
"""

import os
import pykakasi
import argparse
import spacy
import openpyxl
from openpyxl.styles import Font
import pandas as pd
from tqdm import tqdm
from .transliterate import translit
from utils.functions import find_language, set_global_variables

LANGUAGES, NO_LATIN, OBLIGATORY_COLUMNS = set_global_variables()


class Transliterator:
    """
    Class for transliterating sentences from one script to another.

    Attributes:
    -----------
    input_dir : str
        Directory containing files to transliterate.
    language : str
        Source language name.
    instruction : str
        Type of processing ("sentences" or "corrected_transcription").
    device : str
        Device for processing (e.g., "cpu" or "gpu").
    language_code : str
        Language code determined by the find_language function.
    """

    def __init__(self, input_dir, language, instruction, device):
        """
        Initialize the Transliterator with a directory, language, instruction, and device.

        Parameters:
        -----------
        input_dir : str
            Directory containing files to transliterate.
        language : str
            Source language name (will be looked up in LANGUAGES).
        instruction : str
            Type of processing ("sentences" or "corrected_transcription").
        device : str
            Device for processing.
        """
        self.input_dir = input_dir
        self.instruction = instruction
        self.device = device
        self.language_code = find_language(language, LANGUAGES)

    @staticmethod
    def kanji_hiragana_katakana_to_romaji(sentence):
        """
        Convert a Japanese sentence (with Kanji, Hiragana, Katakana) to Romaji.

        Parameters:
        -----------
        sentence : str
            Japanese sentence to convert.

        Returns:
        --------
        romaji_sentence : str
            The sentence converted into Romaji.
        """
        nlp = spacy.load('ja_core_news_trf')
        doc = nlp(sentence)
        kks = pykakasi.kakasi()

        romaji_sentence = ''
        for word in doc:
            morph = word.morph.to_dict()
            if word.text.isascii():
                romaji_word = word.text
            elif word.text == "、":
                romaji_word = ","
            elif word.text == "。":
                romaji_word = "."
            else:
                # Use .get() to safely obtain the reading
                kana_form = morph.get('Reading', str(word))
                romaji_word_lst = kks.convert(kana_form)
                romaji_word = ' '.join([item['hepburn'] for item in romaji_word_lst])
            romaji_sentence += f"{romaji_word} "
        return romaji_sentence.strip()

    def transliterate(self, df):
        """
        Transliterate sentences in the provided DataFrame based on the language code and instruction.

        Parameters:
        -----------
        df : pandas.DataFrame
            DataFrame read from an Excel file.

        Returns:
        --------
        df : pandas.DataFrame
            Updated DataFrame with transliterated sentences.
        """
        if self.instruction == 'sentences':
            source = 'transcription_original_script_utterance_used'
            target = 'latin_transcription_utterance_used'
        elif self.instruction == 'corrected':
            source = 'transcription_original_script'
            target = 'latin_transcription_everything'
        elif self.instruction == 'automatic':
            raise NotImplementedError("Transliteration for automatic transcription is not supported.")
        else:
            raise ValueError(f"Unsupported instruction: {self.instruction}")

        # Initialize target column as empty strings
        df[target] = ""
        df[target] = df[target].astype('object')

        # Process each non-null sentence in the source column
        for sentence in df[source].dropna():
            # Identify all occurrences of the sentence in the DataFrame
            series = df[df.isin([sentence])].stack()
            for idx, _ in series.items():
                # Initialize target cell if it is NaN
                if pd.isna(df.at[idx[0], target]):
                    df.at[idx[0], target] = ""
                # Transliterate based on the language code
                if self.language_code == "ru":
                    transliterated = translit(sentence, 'ru', reversed=True)
                    if transliterated not in df.at[idx[0], target]:
                        df.at[idx[0], target] += f"{transliterated} "
                elif self.language_code == "uk":
                    transliterated = translit(sentence, 'uk', reversed=True)
                    if transliterated not in df.at[idx[0], target]:
                        df.at[idx[0], target] += f"{transliterated} "
                elif self.language_code == "ja":
                    transliterated = Transliterator.kanji_hiragana_katakana_to_romaji(sentence)
                    if transliterated not in df.at[idx[0], target]:
                        df.at[idx[0], target] += f"{transliterated} "
        return df

    def process_data(self):
        """
        Process all Excel files within the input directory and its subdirectories.

        This function searches for files ending with 'annotated.xlsx', processes each file by
        transliterating the content according to the specified instruction, and then saves the file.
        """
        files_to_process = []
        for subdir, _, files in os.walk(self.input_dir):
            for file in files:
                if file.endswith('annotated.xlsx'):
                    files_to_process.append(os.path.join(subdir, file))

        # Process each file individually
        for file_path in tqdm(files_to_process, desc="Processing Files", unit="file"):
            print(f"Processing {file_path}...")
            df = pd.read_excel(file_path)
            df = self.transliterate(df)
            df.to_excel(file_path, index=False)

            # Open the Excel file and modify cell formatting
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active  # Select the first sheet

            # Define the red font style
            red_font = Font(color="FF0000")  # Hex code for red color

            # Define the target column name that should have red font
            if self.instruction == 'corrected':
                target_column = 'latin_transcription_everything'
            elif self.instruction == 'sentences':
                target_column = 'latin_transcription_utterance_used'
            else:
                raise ValueError(f"Unsupported instruction: {self.instruction}")

            # Get the header row (first row) as a tuple of column names
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

            # Ensure the target column exists
            if target_column not in header_row:
                raise ValueError(f"Target column '{target_column}' not found in header row.")

            # Find the column index (1-based) for the target column
            target_index = header_row.index(target_column) + 1

            # Apply red font to each cell in the target column (skip header row)
            for row in ws.iter_rows(min_row=2):
                cell = row[target_index - 1]  # Adjust for 0-based index in row
                if cell.value:
                    cell.font = red_font

            # Save the modified Excel file
            wb.save(file_path)



def main():
    """
    Main function to transliterate transcriptions in all subdirectories.

    Command-line arguments:
    -----------------------
    input_dir : str
        Directory containing files to transliterate.
    instruction : str
        Type of processing ("corrected_transcription" or "sentences").
    source_language : str
        Source language for transliteration.
    """
    parser = argparse.ArgumentParser(description="Automatic transcription")
    parser.add_argument("input_dir", help="Directory with files to transliterate.")
    parser.add_argument("instruction", choices=["corrected_transcription", "sentences"],
                        help="Type of instruction for processing.")
    parser.add_argument("source_language", help="Source language for transliteration.")
    args = parser.parse_args()

    # Verify that the source language is supported
    language_code = None
    for code, name in LANGUAGES.items():
        if name.lower() == args.source_language.lower():
            language_code = code
            print(f"Transliterating for language code: {language_code}")
            break

    if not language_code:
        print(f"Error: Unsupported language '{args.source_language}'.")
        return

    # Create an instance of Transliterator and process the data
    transliterator = Transliterator(args.input_dir, language_code, args.instruction, device="cpu")
    transliterator.process_data()


if __name__ == "__main__":
    main()
